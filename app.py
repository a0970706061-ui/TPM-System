import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from collections import defaultdict
from datetime import datetime
import re

st.title("TPM 保養排程系統")

production_file = st.file_uploader("上傳生產排程 Excel", type=["xlsx"])
maintenance_file = st.file_uploader("上傳保養紀錄 Excel", type=["xlsx"])

selected_month = st.number_input("選擇月份", min_value=1, max_value=12, value=5)
start_day = st.number_input("起始日期", min_value=1, max_value=31, value=11)

exclude_precision = st.checkbox("排除精密加工區", value=True)

max_per_day = st.number_input(
    "每日最多保養台數",
    min_value=1,
    max_value=10,
    value=2
)

def norm_machine(x):
    if x is None:
        return None

    s = str(x).strip().upper()
    s = s.split("\n")[0].strip()
    s = s.replace(" ", "")

    match = re.match(r"^([A-Z]+)-?(\d+)$", s)
    if match:
        prefix = match.group(1)
        num = int(match.group(2))
        return f"{prefix}-{num:03d}"

    return s

def is_machine_code(x):
    if x is None:
        return False

    s = str(x).strip().upper()
    s = s.split("\n")[0].strip()
    s = s.replace(" ", "")

    return bool(re.match(r"^[A-Z]+-?\d+$", s))

def format_date(v):
    if isinstance(v, datetime):
        return f"{v.month}/{v.day}", v

    try:
        dt = pd.to_datetime(v)
        return f"{dt.month}/{dt.day}", dt
    except:
        return str(v), None

def is_precision_area(machine, area):
    machine = str(machine)
    area = str(area)

    if area == "精密加工區":
        return True

    if machine.startswith(("HC-", "HW-", "HMA-", "HLR-")):
        return True

    return False

if st.button("生成TPM排程"):

    if production_file is None:
        st.error("請先上傳生產排程 Excel")

    elif maintenance_file is None:
        st.error("請先上傳保養紀錄 Excel")

    else:
        st.success("開始生成排程")

        # =========================
        # 1. 讀取生產排程
        # =========================
        wb = load_workbook(production_file, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # 只抓指定月份 + 指定起始日之後的日期欄
        date_cols = {}

        for col in range(5, ws.max_column + 1):
            v = ws.cell(1, col).value

            if v is None:
                continue

            date_text, date_dt = format_date(v)

            if (
                date_dt is not None
                and date_dt.month == selected_month
                and date_dt.day >= start_day
            ):
                date_cols[col] = {
                    "顯示日期": date_text,
                    "日期排序": date_dt
                }

        if not date_cols:
            st.error(f"找不到 {selected_month} 月 {start_day} 號之後的日期欄位")
            st.stop()

        # =========================
        # 2. 找每台機台區塊與所有料號列
        # =========================
        machine_info = {}
        machine_part_rows = defaultdict(list)

        current_machine = None
        current_area = None

        for row in range(2, ws.max_row + 1):
            machine_cell = ws.cell(row, 1).value
            area_cell = ws.cell(row, 3).value
            label_cell = ws.cell(row, 4).value

            # 只認真正的機台代號，例如 LA-001、LB-026、LS-098
            if is_machine_code(machine_cell):
                current_machine = norm_machine(machine_cell)
                current_area = str(area_cell).strip() if area_cell is not None else ""
                machine_info[current_machine] = current_area

            # 同一機台區塊內，D欄是「料號」的列都納入
            if current_machine and str(label_cell).strip() == "料號":
                machine_part_rows[current_machine].append(row)

        if not machine_part_rows:
            st.error("沒有抓到任何機台料號列，請確認D欄是否有『料號』")
            st.stop()

        # =========================
        # 3. 判斷每日空機
        # 只要任一料號列有內容，就算生產
        # =========================
        idle_rows = []

        for machine, part_rows in machine_part_rows.items():
            area = machine_info.get(machine, "")

            for col, info in date_cols.items():

                has_schedule = False

                for r in part_rows:
                    val = ws.cell(r, col).value

                    if val is not None and str(val).strip() != "":
                        has_schedule = True
                        break

                if not has_schedule:
                    idle_rows.append({
                        "日期": info["顯示日期"],
                        "日期排序": info["日期排序"],
                        "機台": machine,
                        "廠區": area
                    })

        idle_df = pd.DataFrame(idle_rows)

        if idle_df.empty:
            st.warning("沒有抓到空機資料")
            st.stop()

        # =========================
        # 4. 讀取「所有機台」的上次保養時間
        # =========================
        try:
            maint_df = pd.read_excel(maintenance_file, sheet_name="所有機台")
        except:
            st.error("保養紀錄檔找不到『所有機台』工作表")
            st.stop()

        maint_machine_col = maint_df.columns[0]
        maint_time_col = maint_df.columns[1]

        maintenance_map = {}

        for _, row in maint_df.iterrows():
            machine = norm_machine(row[maint_machine_col])
            last_time = row[maint_time_col]

            if machine is None:
                continue

            if pd.isna(last_time) or str(last_time).strip() in ["無紀錄", "無", "N/A", ""]:
                maintenance_map[machine] = "無紀錄"
            else:
                maintenance_map[machine] = str(last_time)

        idle_df["上次保養時間"] = idle_df["機台"].map(maintenance_map).fillna("無紀錄")

        idle_df["是否無保養紀錄"] = idle_df["上次保養時間"].apply(
            lambda x: "是" if str(x).strip() == "無紀錄" else "否"
        )

        # =========================
        # 5. 讀取總加工時間
        # =========================
        total_time_map = {}

        try:
            time_df = pd.read_excel(maintenance_file, sheet_name="總加工時間")

            time_machine_col = time_df.columns[0]
            time_sec_col = time_df.columns[1]

            for _, row in time_df.iterrows():
                machine = norm_machine(row[time_machine_col])
                sec = row[time_sec_col]

                if machine is None:
                    continue

                try:
                    total_time_map[machine] = float(sec)
                except:
                    total_time_map[machine] = 0

        except:
            st.warning("找不到『總加工時間』工作表，總加工時間先以0計算")

        idle_df["總加工秒數"] = idle_df["機台"].map(total_time_map).fillna(0)

        if idle_df["總加工秒數"].max() > 0:
            threshold = idle_df["總加工秒數"].quantile(0.8)
        else:
            threshold = 0

        idle_df["是否高加工時間"] = idle_df["總加工秒數"].apply(
            lambda x: "是" if x >= threshold and x > 0 else "否"
        )

        # =========================
        # 6. 排除精密加工區
        # =========================
        if exclude_precision:
            idle_df = idle_df[
                ~idle_df.apply(
                    lambda r: is_precision_area(r["機台"], r["廠區"]),
                    axis=1
                )
            ]

        if idle_df.empty:
            st.warning("排除精密加工區後，沒有可安排保養的空機")
            st.stop()

        # =========================
        # 7. 優先原因與排序
        # =========================
        def get_priority(row):
            no_record = row["是否無保養紀錄"] == "是"
            high_time = row["是否高加工時間"] == "是"

            if no_record and high_time:
                return "A"
            elif no_record:
                return "B"
            elif high_time:
                return "C"
            else:
                return "D"

        def get_reason(row):
            if row["優先級"] == "A":
                return "空機 + 無保養紀錄 + 總加工時間高"
            elif row["優先級"] == "B":
                return "空機 + 無保養紀錄"
            elif row["優先級"] == "C":
                return "空機 + 總加工時間高"
            else:
                return "一般空機"

        idle_df["優先級"] = idle_df.apply(get_priority, axis=1)
        idle_df["優先原因"] = idle_df.apply(get_reason, axis=1)

        priority_order = {
            "A": 1,
            "B": 2,
            "C": 3,
            "D": 4
        }

        idle_df["排序"] = idle_df["優先級"].map(priority_order)

        idle_df = idle_df.sort_values(
            ["日期排序", "排序", "總加工秒數"],
            ascending=[True, True, False]
        )

        # =========================
        # 8. 生成保養排程
        # 每天最多 max_per_day
        # 同一台不重複
        # 不跨日補位
        # =========================
        scheduled = set()
        schedule_rows = []

        for date in idle_df["日期"].drop_duplicates():

            day_df = idle_df[idle_df["日期"] == date]
            count = 0

            for _, row in day_df.iterrows():
                machine = row["機台"]

                if machine in scheduled:
                    continue

                schedule_rows.append({
                    "日期": row["日期"],
                    "順序": count + 1,
                    "機台": machine,
                    "優先原因": row["優先原因"],
                    "上次保養時間": row["上次保養時間"],
                    "總加工秒數": int(row["總加工秒數"]),
                    "完成確認": "□"
                })

                scheduled.add(machine)
                count += 1

                if count >= max_per_day:
                    break

        result = pd.DataFrame(schedule_rows)

        if result.empty:
            st.warning("沒有產生任何保養排程")
            st.stop()

        # =========================
        # 9. 顯示結果
        # =========================
        st.write("### TPM保養排程")
        st.dataframe(result)

        st.write("### 每日空機優先清單")
        display_idle_df = idle_df.drop(columns=["日期排序", "排序", "廠區", "優先級"])
        st.dataframe(display_idle_df)

        # =========================
        # 10. 下載 Excel
        # =========================
        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            result.to_excel(writer, index=False, sheet_name="TPM保養排程")
            display_idle_df.to_excel(writer, index=False, sheet_name="每日空機優先清單")

        st.download_button(
            label="下載 TPM保養排程 Excel",
            data=output.getvalue(),
            file_name=f"{selected_month}月_TPM保養排程.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )